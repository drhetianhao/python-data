import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook, Workbook

data = pd.read_excel(r"D:\test.xlsx")
data = data.drop(
    columns=[
        "dtime",
        "packet_id",
        "packet_size",
        "cum_size",
        "delta_duration",
        "location_id",
    ]
)
only10folddata = data[data["cum_packets_recv"] % 10 == 0]
datagroup = only10folddata.groupby("batch")
# df = datagroup.get_group(1)
# print(datagroup.get_group(1))

filename = r"D:\my_excel_file.xlsx"
if os.path.isfile(filename):
    os.remove(filename)
    book = Workbook()
    sheet = book.active
    sheet["A1"] = "This is a test"
    book.save(filename)
    print("File exist")
else:
    book = Workbook()
    sheet = book.active
    book.save(filename)
    print("File not exist")

for name, group in datagroup:
    print(name)
    df = group
    top_row = pd.DataFrame({"cum_packets_recv": [0], "cum_duration": [0]})
    df = pd.concat([top_row, df]).reset_index(drop=True)
    df["cum_duration"] = pd.to_numeric(df["cum_duration"])
    difference = df["cum_duration"]
    difference = difference.diff(periods=3)
    df1 = difference.dropna()
    df2 = df1.quantile(0.7)
    df1 = df1.to_numpy()
    mbps = 1400 * 8 * 30 / df2
    df1 = np.append(df1, df2)
    df1 = np.append(df1, mbps)
    df1 = df1[::-1]
    df = pd.DataFrame(df1).transpose()

    try:
        writer = pd.ExcelWriter(filename, mode="a", engine="openpyxl")
        writer.book = load_workbook(filename)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, startrow=name, index=False, header=False)
        writer.save()
    except:
        print("file not there")
